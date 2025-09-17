from flask import (
    Flask,
    render_template,
    request,
    jsonify,
    redirect,
    url_for,
    flash,
    send_file,
    make_response,
)
from flask_socketio import SocketIO, emit
import os
import json
from datetime import datetime
from werkzeug.utils import secure_filename
import threading
import time
from io import BytesIO, StringIO
import tempfile

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("Warning: pandas not available, using alternative data processing")
from agents.agent1 import Agent1
from agents.agent2 import Agent2
from agents.reflect_agent import ReflectAgent
from sample_data import get_sample_campaign_data, get_sample_customer_data
from utils.export_utils import ExportUtils

try:
    from langgraph_workflow import run_langgraph_workflow
except ImportError:

    def run_langgraph_workflow(*args, **kwargs):
        return {"error": "LangGraph workflow not available"}


app = Flask(__name__)
app.secret_key = "your-secret-key-here"  # Change this in production
app.config["UPLOAD_FOLDER"] = "uploads"
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100MB max file size
socketio = SocketIO(app, cors_allowed_origins="*")

# Ensure upload directory exists
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

# Allowed file extensions
ALLOWED_EXTENSIONS = {"csv", "json", "xlsx", "xls"}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route("/")
def landing_page():
    """Landing page with project overview and navigation"""
    return render_template("landing.html")


@app.route("/dashboard")
def dashboard():
    """Main dashboard for data upload and agent processing"""
    return render_template("dashboard.html")


@app.route("/upload", methods=["POST"])
def upload_file():
    """Handle file upload and data processing"""
    try:
        if "file" not in request.files:
            return jsonify({"success": False, "error": "No file selected"})

        files = request.files.getlist("file")
        if not files or files[0].filename == "":
            return jsonify({"success": False, "error": "No file selected"})

        uploaded_files = []
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file.save(filepath)

                try:
                    # Process the uploaded file
                    data = process_uploaded_file(filepath)
                    uploaded_files.append(
                        {
                            "filename": filename,
                            "size": len(data) if isinstance(data, list) else 1,
                            "preview": (
                                data[:3] if isinstance(
                                    data, list) else str(data)[:100]
                            ),
                        }
                    )
                except Exception as e:
                    return jsonify(
                        {
                            "success": False,
                            "error": f"Error processing {filename}: {str(e)}",
                        }
                    )
            else:
                return jsonify(
                    {"success": False, "error": f"Invalid file type: {file.filename}"}
                )

        return jsonify(
            {
                "success": True,
                "message": f"Successfully uploaded {len(uploaded_files)} file(s)",
                "files": uploaded_files,
            }
        )

    except Exception as e:
        return jsonify({"success": False, "error": f"Upload error: {str(e)}"})


def process_agents_async(data_type="sample"):
    """Async processing with real-time status updates"""
    global current_agent1, stored_segments
    try:
        socketio.emit(
            "agent_status", {"status": "starting",
                             "message": "Initializing agents..."}
        )

        if data_type == "sample":
            campaigns = get_sample_campaign_data()
            customers = get_sample_customer_data()
        else:
            campaigns = get_sample_campaign_data()
            customers = get_sample_customer_data()

        # Initialize agents
        socketio.emit(
            "agent_status",
            {"status": "initializing", "message": "Setting up Agent 1..."},
        )
        agent1 = Agent1()
        current_agent1 = agent1  # Set global reference for SocketIO handlers

        # Process through Agent1 with status updates (will pause for goal approval)
        socketio.emit(
            "agent_status",
            {
                "status": "agent1_thinking",
                "message": "Agent 1 analyzing customer data...",
                "step": "data_analysis",
            },
        )

        agent1_result = agent1.run_complete_workflow_with_status(
            customers, socketio)

        # Check if workflow is paused for approval
        if agent1_result.get("status") == "awaiting_goal_approval":
            # Store segments for later use when continuing workflow
            from agents.agent1 import CustomerSegment

            stored_segments = [
                CustomerSegment(**seg) for seg in agent1_result.get("segments", [])
            ]

            socketio.emit(
                "agent_status",
                {
                    "status": "awaiting_goal_approval",
                    "message": "Agent 1 generated business goals - awaiting human approval",
                    "result": agent1_result,
                },
            )

            # Workflow will continue when goals are approved via SocketIO
            print(
                "🔄 Workflow paused for business goal approval. Waiting for user input..."
            )

        elif agent1_result.get("status") == "failed":
            socketio.emit(
                "agent_status",
                {
                    "status": "error",
                    "message": f'Agent 1 failed: {agent1_result.get("error", "Unknown error")}',
                },
            )
        else:
            # This shouldn't happen with the new workflow, but keeping as fallback
            socketio.emit(
                "agent_status",
                {
                    "status": "agent1_complete",
                    "message": "Agent 1 completed analysis",
                    "result": agent1_result,
                },
            )

    except Exception as e:
        socketio.emit(
            "agent_status",
            {"status": "error", "message": f"Processing failed: {str(e)}"},
        )


@app.route("/process_agents", methods=["POST"])
def process_agents():
    """Start agent processing in background thread"""
    try:
        data_type = request.json.get("data_type", "sample")

        # Start processing in background thread
        thread = threading.Thread(
            target=process_agents_async, args=(data_type,))
        thread.daemon = True
        thread.start()

        return jsonify(
            {"success": True, "message": "Processing started, check real-time status"}
        )

    except Exception as e:
        return jsonify(
            {"success": False, "error": str(
                e), "timestamp": datetime.now().isoformat()}
        )


@app.route("/api/status")
def get_status():
    """Get system status and health check"""
    try:
        # Test agent initialization
        agent1 = Agent1()
        agent2 = Agent2()

        return jsonify(
            {
                "status": "healthy",
                "agents": {"agent1": "ready", "agent2": "ready"},
                "timestamp": datetime.now().isoformat(),
            }
        )
    except Exception as e:
        return jsonify(
            {
                "status": "error",
                "error": str(e),
                "timestamp": datetime.now().isoformat(),
            }
        )


def process_uploaded_file(filepath):
    """Process uploaded file and return data"""
    file_ext = filepath.rsplit(".", 1)[1].lower()

    if file_ext == "csv":
        if PANDAS_AVAILABLE:
            df = pd.read_csv(filepath)
            return df.to_dict("records")
        else:
            # Simple CSV parsing without pandas
            import csv

            with open(filepath, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                return list(reader)
    elif file_ext == "json":
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    elif file_ext in ["xlsx", "xls"]:
        if PANDAS_AVAILABLE:
            df = pd.read_excel(filepath)
            return df.to_dict("records")
        else:
            # Use openpyxl for Excel files without pandas
            from openpyxl import load_workbook

            wb = load_workbook(filepath)
            ws = wb.active
            data = []
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append(dict(zip(headers, row)))
            return data
    else:
        raise ValueError(f"Unsupported file type: {file_ext}")


# Global agent instance for state management
current_agent1 = None
campaign_approval_received = False
approved_campaigns = []

# Global Reflect Agent instance
reflect_agent = ReflectAgent()


def continue_agent1_workflow():
    """Continue Agent1 workflow after goal approval"""
    global current_agent1, stored_segments
    try:
        if (
            current_agent1
            and hasattr(current_agent1, "approved_goals")
            and hasattr(current_agent1, "continue_workflow_after_approval")
        ):
            socketio.emit(
                "agent_status",
                {
                    "status": "agent1_continuing",
                    "message": "Continuing Agent 1 workflow after goal approval...",
                },
            )

            # Continue the workflow with approved goals
            agent1_result = current_agent1.continue_workflow_after_approval(
                stored_segments, current_agent1.approved_goals, socketio
            )

            # Proceed to Agent2 after Agent1 completes
            if agent1_result.get("status") != "failed":
                socketio.emit(
                    "agent_status",
                    {
                        "status": "agent1_complete",
                        "message": "Agent 1 completed analysis",
                        "result": agent1_result,
                    },
                )

                # Start Agent2
                from agents.agent2 import Agent2

                agent2 = Agent2()

                socketio.emit(
                    "agent_status",
                    {
                        "status": "agent2_starting",
                        "message": "Starting Agent 2 implementation workflow...",
                        "step": "agent2_initialization",
                    },
                )

                agent2_result = agent2.run_implementation_workflow_with_status(
                    agent1_result, socketio
                )

                socketio.emit(
                    "agent_status",
                    {
                        "status": "completed",
                        "message": "All agents completed successfully!",
                        "agent1_result": agent1_result,
                        "agent2_result": agent2_result,
                        "timestamp": datetime.now().isoformat(),
                    },
                )

        else:
            socketio.emit(
                "agent_status",
                {
                    "status": "error",
                    "message": "Unable to continue workflow - agent or goals not available",
                },
            )
    except Exception as e:
        socketio.emit(
            "agent_status",
            {"status": "error",
                "message": f"Error continuing workflow: {str(e)}"},
        )


# SocketIO event handlers for goal management
@socketio.on("approve_selected_goals")
def handle_approve_selected_goals(data):
    """Handle approval of selected business goals and continue workflow"""
    global current_agent1
    goal_indices = data.get("goal_indices", [])

    if current_agent1 and hasattr(current_agent1, "pending_goals"):
        # Update approved goals based on selected indices
        current_agent1.approved_goals = []
        for idx in goal_indices:
            if 0 <= idx < len(current_agent1.pending_goals):
                current_agent1.approved_goals.append(
                    current_agent1.pending_goals[idx])

        current_agent1.business_goals = current_agent1.approved_goals
        print(f"✅ Approved {len(current_agent1.approved_goals)} goals via UI")

    # Emit approval success with animation trigger
    emit(
        "goals_approval_success",
        {
            "message": f"Successfully approved {len(goal_indices)} business goals!",
            "approved_count": (
                len(current_agent1.approved_goals)
                if current_agent1 and hasattr(current_agent1, "approved_goals")
                else 0
            ),
            "show_animation": True,
        },
    )

    # Continue the workflow after approval
    socketio.start_background_task(continue_agent1_workflow)

    print(f"Goals approved: {goal_indices}. Continuing workflow...")


@socketio.on("modify_specific_goal")
def handle_modify_specific_goal(data):
    global current_agent1
    goal_index = data.get("goal_index")
    field = data.get("field")
    new_value = data.get("new_value")
    feedback = data.get("feedback")  # For general feedback-based modifications

    if current_agent1 and hasattr(current_agent1, "pending_goals"):
        if 0 <= goal_index < len(current_agent1.pending_goals):
            goal = current_agent1.pending_goals[goal_index]

            # Handle specific field modifications
            if field and new_value is not None:
                if field == "name":
                    goal.name = new_value
                elif field == "description":
                    goal.description = new_value
                elif field == "priority":
                    goal.priority = new_value.lower()
                elif field == "target_segments":
                    goal.target_segments = [
                        s.strip() for s in new_value.split(",") if s.strip()
                    ]
                elif field == "success_metrics":
                    goal.success_metrics = [
                        s.strip() for s in new_value.split(",") if s.strip()
                    ]
                elif field == "timeline":
                    goal.timeline = new_value

                print(
                    f"✅ Modified goal {goal_index + 1}: {field} = {new_value}")

                # Send updated goal data back to frontend
                emit(
                    "goal_updated",
                    {
                        "goal_index": goal_index,
                        "updated_goal": {
                            "name": goal.name,
                            "description": goal.description,
                            "priority": goal.priority,
                            "target_segments": goal.target_segments,
                            "success_metrics": goal.success_metrics,
                            "timeline": getattr(goal, "timeline", ""),
                        },
                    },
                )

            # Handle general feedback-based modifications (legacy support)
            elif feedback:
                print(f"📝 Feedback for goal {goal_index + 1}: {feedback}")
                emit(
                    "modification_feedback_received",
                    {"goal_index": goal_index, "feedback": feedback},
                )

    emit(
        "modification_completed",
        {
            "message": f"Goal {goal_index + 1} modified successfully",
            "goal_index": goal_index,
            "field": field,
            "new_value": new_value,
        },
    )


@socketio.on("reject_goal")
def handle_reject_goal(data):
    global current_agent1
    goal_index = data.get("goal_index")

    if current_agent1 and hasattr(current_agent1, "pending_goals"):
        if 0 <= goal_index < len(current_agent1.pending_goals):
            rejected_goal = current_agent1.pending_goals.pop(goal_index)
            print(f"❌ Rejected goal: {rejected_goal.name}")

    emit(
        "goal_rejected",
        {"message": f"Goal {goal_index + 1} rejected", "goal_index": goal_index},
    )


@socketio.on("delete_goal")
def handle_delete_goal(data):
    global current_agent1
    goal_index = data.get("goal_index")

    if current_agent1 and hasattr(current_agent1, "pending_goals"):
        if 0 <= goal_index < len(current_agent1.pending_goals):
            deleted_goal = current_agent1.pending_goals.pop(goal_index)
            print(f"🗑️ Deleted goal: {deleted_goal.name}")

            emit(
                "goal_deleted",
                {
                    "goal_index": goal_index,
                    "message": f'Goal "{deleted_goal.name}" has been deleted successfully',
                },
            )
        else:
            emit("error", {"message": "Invalid goal index"})
    else:
        emit("error", {"message": "No goals available to delete"})


@socketio.on("add_custom_goal")
def handle_add_custom_goal(data):
    global current_agent1
    from agents.agent1 import BusinessGoal

    name = data.get("name", "")
    description = data.get("description", "")
    priority = data.get("priority", "medium").lower()
    target_segments = [
        s.strip() for s in data.get("target_segments", "").split(",") if s.strip()
    ]
    success_metrics = [
        m.strip() for m in data.get("success_metrics", "").split(",") if m.strip()
    ]

    if current_agent1 and hasattr(current_agent1, "pending_goals"):
        custom_goal = BusinessGoal(
            name=name,
            description=description,
            target_segments=target_segments,
            success_metrics=success_metrics,
            priority=priority,
        )
        current_agent1.pending_goals.append(custom_goal)
        print(f"✅ Added custom goal: {name}")

    emit(
        "custom_goal_added",
        {
            "message": f"Custom goal added: {name}",
            "goal": {
                "name": name,
                "description": description,
                "priority": priority,
                "target_segments": target_segments,
                "success_metrics": success_metrics,
            },
        },
    )


@socketio.on("finish_goal_selection")
def handle_finish_goal_selection():
    global current_agent1

    if current_agent1:
        if (
            not hasattr(current_agent1, "approved_goals")
            or not current_agent1.approved_goals
        ):
            # If no goals were explicitly approved, use all pending goals
            if hasattr(current_agent1, "pending_goals"):
                current_agent1.approved_goals = current_agent1.pending_goals.copy()
                current_agent1.business_goals = current_agent1.approved_goals
                print(
                    f"⚠️ No goals explicitly approved. Using all {len(current_agent1.approved_goals)} goals."
                )

        print(
            f"✅ Goal selection completed. Final count: {len(current_agent1.approved_goals)} goals"
        )

    emit(
        "goal_selection_completed",
        {
            "message": "Goal selection completed. Proceeding to campaign generation...",
            "approved_count": (
                len(current_agent1.approved_goals)
                if current_agent1 and hasattr(current_agent1, "approved_goals")
                else 0
            ),
        },
    )
    print("Goal selection finished")


@socketio.on("approve_goals")
def handle_approve_goals():
    """Handle approval of all business goals and continue workflow"""
    global current_agent1

    # Approve all pending goals
    if current_agent1 and hasattr(current_agent1, "pending_goals"):
        current_agent1.approved_goals = current_agent1.pending_goals.copy()
        current_agent1.business_goals = current_agent1.approved_goals

        # Emit approval success with animation trigger
        emit(
            "goals_approval_success",
            {
                "message": f"Successfully approved all {len(current_agent1.approved_goals)} business goals!",
                "approved_count": len(current_agent1.approved_goals),
                "show_animation": True,
            },
        )

        # Continue the workflow after approval
        socketio.start_background_task(continue_agent1_workflow)

        print(
            f"All {len(current_agent1.approved_goals)} goals approved. Continuing workflow..."
        )
    else:
        emit("approval_received", {
             "message": "No goals available for approval"})
        print("No goals available for approval")


@socketio.on("modify_goals")
def handle_modify_goals(data):
    feedback = data.get("feedback")
    emit("modification_requested", {"feedback": feedback})
    print(f"Goals modification requested: {feedback}")


@socketio.on("approve_campaigns")
def handle_approve_campaigns():
    global campaign_approval_received, approved_campaigns, current_agent1
    campaign_approval_received = True

    # Store approved campaigns from Agent 1
    if current_agent1 and hasattr(current_agent1, "campaigns"):
        approved_campaigns = current_agent1.campaigns.copy()
        print(
            f"✅ Stored {len(approved_campaigns)} approved campaigns for display")

        # Emit campaigns data to frontend for immediate display
        campaigns_data = []
        for campaign in approved_campaigns:
            campaign_dict = {
                "name": campaign.name,
                "description": campaign.description,
                "target_segments": campaign.target_segments,
                "business_goals": campaign.business_goals,
                "channels": campaign.channels,
                "content": campaign.content,
                "timeline": campaign.timeline,
                "success_metrics": campaign.success_metrics,
                "budget_estimate": campaign.budget_estimate,
            }
            campaigns_data.append(campaign_dict)

        # Emit to update the marketing campaigns section
        socketio.emit(
            "campaigns_approved_and_stored",
            {
                "campaigns": campaigns_data,
                "message": f"Successfully approved and stored {len(campaigns_data)} campaigns!",
            },
        )
    else:
        print("⚠️ No campaigns found in Agent 1 to approve")

    emit("approval_received", {"message": "All campaigns approved"})
    print("All campaigns approved")
    # Continue the workflow after approval
    socketio.emit(
        "agent_status",
        {
            "status": "agent1_continuing",
            "message": "Continuing Agent 1 workflow after approval...",
        },
    )


@socketio.on("modify_campaigns")
def handle_modify_campaigns(data):
    feedback = data.get("feedback")
    emit("modification_requested", {"feedback": feedback})
    print(f"Campaigns modification requested: {feedback}")


@app.route("/api/campaigns", methods=["GET"])
def get_approved_campaigns():
    """Get approved campaigns for display in marketing campaigns section"""
    global approved_campaigns

    try:
        campaigns_data = []
        for campaign in approved_campaigns:
            campaign_dict = {
                "name": campaign.name,
                "description": campaign.description,
                "target_segments": campaign.target_segments,
                "business_goals": campaign.business_goals,
                "channels": campaign.channels,
                "content": campaign.content,
                "timeline": campaign.timeline,
                "success_metrics": campaign.success_metrics,
                "budget_estimate": campaign.budget_estimate,
            }
            campaigns_data.append(campaign_dict)

        return jsonify(
            {"success": True, "campaigns": campaigns_data,
                "count": len(campaigns_data)}
        )
    except Exception as e:
        return (
            jsonify(
                {
                    "success": False,
                    "message": f"Error retrieving campaigns: {str(e)}",
                    "campaigns": [],
                    "count": 0,
                }
            ),
            500,
        )


@app.route("/clear_cache", methods=["POST"])
def clear_cache():
    """Clear the context store cache and reset agent memory"""
    try:
        # Clear Agent 2's context store
        agent2 = Agent2()
        if hasattr(agent2, "context_store") and agent2.context_store:
            # Clear the vector store
            agent2.context_store.clear()

        # Clear any global variables
        global current_agent1, campaign_approval_received, approved_campaigns
        current_agent1 = None
        campaign_approval_received = False
        approved_campaigns = []

        return jsonify({"success": True, "message": "Cache cleared successfully"})
    except Exception as e:
        return (
            jsonify(
                {"success": False, "message": f"Error clearing cache: {str(e)}"}),
            500,
        )


# Export endpoints
@app.route("/export/json")
def export_json():
    """Export results as JSON"""
    try:
        global current_agent1
        if not current_agent1:
            return jsonify({"error": "No data available for export"}), 404

        # Prepare export data
        export_data = {
            "timestamp": datetime.now().isoformat(),
            "segments": [
                {
                    "name": seg.name,
                    "description": seg.description,
                    "size": seg.size_estimate or 0,
                    "priority_score": seg.priority_score,
                    "criteria": seg.criteria,
                    "characteristics": seg.characteristics,
                }
                for seg in current_agent1.segments
            ],
            "business_goals": [
                {
                    "name": goal.name,
                    "description": goal.description,
                    "target_segments": goal.target_segments,
                    "success_metrics": goal.success_metrics,
                    "priority": goal.priority,
                    "timeline": goal.timeline,
                }
                for goal in current_agent1.business_goals
            ],
            "campaigns": [
                {
                    "name": camp.name,
                    "description": camp.description,
                    "target_segments": camp.target_segments,
                    "channels": camp.channels,
                    "budget_estimate": camp.budget_estimate,
                    "success_metrics": camp.success_metrics,
                    "optimization_score": camp.optimization_score,
                }
                for camp in current_agent1.campaigns
            ],
        }

        response = make_response(json.dumps(export_data, indent=2))
        response.headers["Content-Type"] = "application/json"
        response.headers["Content-Disposition"] = (
            f'attachment; filename=marketing_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        )
        return response

    except Exception as e:
        return jsonify({"error": f"Export failed: {str(e)}"}), 500


@app.route("/export/excel")
def export_excel():
    """Export results as Excel file"""
    try:
        global current_agent1
        if not current_agent1:
            return jsonify({"error": "No data available for export"}), 404

        if not PANDAS_AVAILABLE:
            return jsonify({"error": "Excel export requires pandas"}), 500

        # Create Excel file in memory
        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Segments sheet
            segments_data = [
                {
                    "Name": seg.name,
                    "Description": seg.description,
                    "Size": seg.size_estimate or 0,
                    "Priority Score": seg.priority_score,
                    "Criteria": str(seg.criteria),
                    "Characteristics": str(seg.characteristics),
                }
                for seg in current_agent1.segments
            ]

            if segments_data:
                pd.DataFrame(segments_data).to_excel(
                    writer, sheet_name="Customer Segments", index=False
                )

            # Goals sheet
            goals_data = [
                {
                    "Name": goal.name,
                    "Description": goal.description,
                    "Target Segments": ", ".join(goal.target_segments),
                    "Success Metrics": ", ".join(goal.success_metrics),
                    "Priority": goal.priority,
                    "Timeline": goal.timeline,
                }
                for goal in current_agent1.business_goals
            ]

            if goals_data:
                pd.DataFrame(goals_data).to_excel(
                    writer, sheet_name="Business Goals", index=False
                )

            # Campaigns sheet
            campaigns_data = [
                {
                    "Name": camp.name,
                    "Description": camp.description,
                    "Target Segments": ", ".join(camp.target_segments),
                    "Channels": ", ".join(camp.channels),
                    "Budget Estimate": camp.budget_estimate,
                    "Success Metrics": ", ".join(camp.success_metrics or []),
                    "Optimization Score": camp.optimization_score,
                }
                for camp in current_agent1.campaigns
            ]

            if campaigns_data:
                pd.DataFrame(campaigns_data).to_excel(
                    writer, sheet_name="Marketing Campaigns", index=False
                )

        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f'marketing_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        return jsonify({"error": f"Excel export failed: {str(e)}"}), 500


@app.route("/export/pdf")
def export_pdf():
    """Export results as PDF file"""
    try:
        global current_agent1
        if not current_agent1:
            return jsonify({"error": "No data available for export"}), 404

        # Try to import reportlab for PDF generation
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import (
                SimpleDocTemplate,
                Paragraph,
                Spacer,
                Table,
                TableStyle,
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
        except ImportError:
            return jsonify({"error": "PDF export requires reportlab package"}), 500

        # Create PDF in memory
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            spaceAfter=30,
            alignment=1,  # Center alignment
        )
        story.append(
            Paragraph("Marketing Intelligence Analysis Report", title_style))
        story.append(Spacer(1, 20))

        # Timestamp
        story.append(
            Paragraph(
                f'Generated on: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}',
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 20))

        # Customer Segments Section
        story.append(Paragraph("Customer Segments", styles["Heading2"]))
        story.append(Spacer(1, 12))

        for i, seg in enumerate(current_agent1.segments, 1):
            story.append(Paragraph(f"{i}. {seg.name}", styles["Heading3"]))
            story.append(
                Paragraph(
                    f"<b>Description:</b> {seg.description}", styles["Normal"])
            )
            story.append(
                Paragraph(
                    f'<b>Estimated Size:</b> {seg.size_estimate or "Not calculated"} customers',
                    styles["Normal"],
                )
            )
            story.append(
                Paragraph(
                    f"<b>Priority Score:</b> {seg.priority_score}/10", styles["Normal"]
                )
            )
            story.append(Spacer(1, 12))

        # Business Goals Section
        story.append(Paragraph("Business Goals", styles["Heading2"]))
        story.append(Spacer(1, 12))

        for i, goal in enumerate(current_agent1.business_goals, 1):
            story.append(Paragraph(f"{i}. {goal.name}", styles["Heading3"]))
            story.append(
                Paragraph(
                    f"<b>Description:</b> {goal.description}", styles["Normal"])
            )
            story.append(
                Paragraph(
                    f"<b>Priority:</b> {goal.priority.title()}", styles["Normal"])
            )
            story.append(
                Paragraph(
                    f'<b>Target Segments:</b> {", ".join(goal.target_segments)}',
                    styles["Normal"],
                )
            )
            story.append(Spacer(1, 12))

        # Marketing Campaigns Section
        story.append(Paragraph("Marketing Campaigns", styles["Heading2"]))
        story.append(Spacer(1, 12))

        for i, camp in enumerate(current_agent1.campaigns, 1):
            story.append(Paragraph(f"{i}. {camp.name}", styles["Heading3"]))
            story.append(
                Paragraph(
                    f"<b>Description:</b> {camp.description}", styles["Normal"])
            )
            story.append(
                Paragraph(
                    f'<b>Channels:</b> {", ".join(camp.channels)}', styles["Normal"]
                )
            )
            story.append(
                Paragraph(
                    f'<b>Target Segments:</b> {", ".join(camp.target_segments)}',
                    styles["Normal"],
                )
            )
            if camp.optimization_score:
                story.append(
                    Paragraph(
                        f"<b>Optimization Score:</b> {camp.optimization_score}%",
                        styles["Normal"],
                    )
                )
            story.append(Spacer(1, 12))

        # Build PDF
        doc.build(story)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'marketing_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
            mimetype="application/pdf",
        )

    except Exception as e:
        return jsonify({"error": f"PDF export failed: {str(e)}"}), 500


# Implementation Plan Export Routes
export_utils = ExportUtils()


@app.route("/export/implementation/json")
def export_implementation_json():
    """Export implementation plan as JSON"""
    try:
        # Get implementation data from session or use sample data
        implementation_data = (
            session.get("implementation_plan")
            if "session" in globals()
            else export_utils.get_sample_data()
        )

        buffer, filename = export_utils.export_to_json(implementation_data)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/json",
        )

    except Exception as e:
        return jsonify({"error": f"Implementation JSON export failed: {str(e)}"}), 500


@app.route("/export/implementation/excel")
def export_implementation_excel():
    """Export implementation plan as Excel"""
    try:
        # Get implementation data from session or use sample data
        implementation_data = (
            session.get("implementation_plan")
            if "session" in globals()
            else export_utils.get_sample_data()
        )

        buffer, filename = export_utils.export_to_excel(implementation_data)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        return jsonify({"error": f"Implementation Excel export failed: {str(e)}"}), 500


@app.route("/export/implementation/pdf")
def export_implementation_pdf():
    """Export implementation plan as PDF"""
    try:
        # Get implementation data from session or use sample data
        implementation_data = (
            session.get("implementation_plan")
            if "session" in globals()
            else export_utils.get_sample_data()
        )

        buffer, filename = export_utils.export_to_pdf(implementation_data)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/pdf",
        )

    except Exception as e:
        return jsonify({"error": f"Implementation PDF export failed: {str(e)}"}), 500


# Reflect Agent Chat Routes
@app.route("/chat")
def chat_interface():
    """Render chat interface"""
    return render_template("chat.html")


@app.route("/api/chat/start", methods=["POST"])
def start_chat_session():
    """Start a new chat session"""
    try:
        session_id = reflect_agent.start_session()
        return jsonify(
            {
                "success": True,
                "session_id": session_id,
                "message": "Chat session started successfully",
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/chat/message", methods=["POST"])
def send_chat_message():
    """Send message to Reflect Agent"""
    try:
        data = request.get_json()
        message = data.get("message", "")
        context = data.get("context", {})

        if not message.strip():
            return jsonify({"success": False, "error": "Message cannot be empty"})

        # Process message through Reflect Agent
        result = reflect_agent.process_message(message, context)

        return jsonify({"success": True, "result": result})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/chat/history", methods=["GET"])
def get_chat_history():
    """Get chat history for current session"""
    try:
        history = reflect_agent.get_chat_history()
        return jsonify(
            {
                "success": True,
                "history": history,
                "session_summary": reflect_agent.get_session_summary(),
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/chat/memory", methods=["GET"])
def get_memory_context():
    """Get memory context"""
    try:
        query = request.args.get("query", "")
        context = reflect_agent.get_memory_context(query if query else None)
        return jsonify({"success": True, "context": context})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/chat/action", methods=["POST"])
def trigger_system_action():
    """Trigger system actions from chat"""
    try:
        data = request.get_json()
        action = data.get("action", "")
        parameters = data.get("parameters", {})

        result = reflect_agent.trigger_system_action(action, parameters)

        return jsonify({"success": True, "result": result})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/chat/clear", methods=["POST"])
def clear_chat_session():
    """Clear current chat session"""
    try:
        reflect_agent.clear_session()
        return jsonify({"success": True, "message": "Chat session cleared"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# SocketIO handlers for real-time chat
@socketio.on("join_chat")
def handle_join_chat(data):
    """Handle user joining chat"""
    try:
        session_id = data.get("session_id")
        if not session_id:
            session_id = reflect_agent.start_session()

        emit(
            "chat_joined",
            {"success": True, "session_id": session_id,
                "message": "Connected to chat"},
        )

        # Send chat history if available
        history = reflect_agent.get_chat_history()
        if history:
            emit(
                "chat_history",
                {
                    "history": history,
                    "session_summary": reflect_agent.get_session_summary(),
                },
            )

    except Exception as e:
        emit("chat_error", {"error": str(e), "message": "Failed to join chat"})


@socketio.on("send_message")
def handle_send_message(data):
    """Handle incoming chat message"""
    try:
        message = data.get("message", "").strip()
        context = data.get("context", {})

        if not message:
            emit("chat_error", {"error": "Message cannot be empty"})
            return

        # Emit typing indicator
        emit("agent_typing", {"typing": True})

        # Process message through Reflect Agent
        result = reflect_agent.process_message(message, context)

        # Stop typing indicator
        emit("agent_typing", {"typing": False})

        # Send response
        emit("message_response", {"success": True, "result": result})

    except Exception as e:
        emit("agent_typing", {"typing": False})
        emit("chat_error", {"error": str(e),
             "message": "Failed to process message"})


@socketio.on("trigger_action")
def handle_trigger_action(data):
    """Handle system action triggers from chat"""
    try:
        action = data.get("action", "")
        parameters = data.get("parameters", {})

        # Emit action started
        emit("action_started", {"action": action,
             "message": f"Starting {action}..."})

        # Trigger action
        result = reflect_agent.trigger_system_action(action, parameters)

        # Emit action result
        emit("action_completed", {"action": action, "result": result})

    except Exception as e:
        emit("action_error", {"action": data.get(
            "action", "unknown"), "error": str(e)})


@socketio.on("get_suggestions")
def handle_get_suggestions(data):
    """Handle request for contextual suggestions"""
    try:
        query = data.get("query", "")
        context = reflect_agent.get_memory_context(query if query else None)

        # Generate suggestions based on context
        suggestions = [
            {
                "title": "Campaign Analysis",
                "description": "Analyze your campaign performance metrics",
                "action": "analyze_campaign",
            },
            {
                "title": "Customer Insights",
                "description": "Get detailed customer segmentation analysis",
                "action": "analyze_customers",
            },
            {
                "title": "Strategy Optimization",
                "description": "Optimize your marketing strategy based on data",
                "action": "optimize_strategy",
            },
            {
                "title": "Export Report",
                "description": "Generate and export comprehensive report",
                "action": "export_report",
            },
        ]

        emit("suggestions_response", {
             "suggestions": suggestions, "context": context})

    except Exception as e:
        emit("suggestions_error", {"error": str(e)})


@socketio.on("clear_chat")
def handle_clear_chat():
    """Handle chat clearing"""
    try:
        reflect_agent.clear_session()
        emit("chat_cleared", {"success": True,
             "message": "Chat session cleared"})
    except Exception as e:
        emit("chat_error", {"error": str(e),
             "message": "Failed to clear chat"})


if __name__ == "__main__":
    socketio.run(app, debug=True, host="0.0.0.0", port=5000)
